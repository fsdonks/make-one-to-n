#!/usr/bin/env python
# coding: utf-8

# In[ ]:


import pandas as pd
from functools import partial
from pathlib import Path
import openpyxl
import sys
# # TAA Post Processing

# ## Output Checking

# ### Standard Capacity Analysis Run with Default Initial Conditions


# In order to plot a line chart of TotalRequired and Deployed, we group by time and sum the values so that we have the total TotalRequired and Deployed for each day.  If you don't reset_index, you get a multi-index dataframe from groupby, which you can't plot, but functions called on groupby (like sum() here) will sum the values in each group.

# dtrends = root+ "DemandTrends.txt"
# df=pd.read_csv(dtrends, sep='\t')
# df.head()
# group_df = df.groupby(['t']).sum().reset_index()
# group_df.head()

# plt.plot('t', 'TotalRequired', data=group_df)
# plt.plot('t', 'Deployed', data=group_df)

# ### Random Initial Conditions Output Checks

# We've been storing the results the the parent directory alongside the MARATHON workbook.  results.txt is from random initial condition runs from marathon.analysis.random.

# In[ ]:

def check_rand_results():
    #resources location
    resources="/home/craig/workspace/taa_processor/resources/"
    results = resources+ "results.txt" 
    results
    
    
    # In[ ]:
    
    
    df=pd.read_csv(results, sep='\t')
    df.head()
    
    
    # Here we count the number records for each \[SRC, AC\] group.  For x initial condition reps and y phases, we should have x*y records.  This is essentially pivoting in Python by count.
    
    # In[ ]:
    
    
    group_df = df.groupby(by=['SRC', 'AC']).count().reset_index()
    group_df.head()
    
    
    # Check for any \[SRC, AC\] tuple that doesn't have x*y records.
    
    # In[ ]:
    
    
    group_df[group_df['rep-seed']!=12]


# ## Post Processing

# We'd like to compute Score and Excess for each \[SRC, AC\] tuple.  
# 
# First, average NG fill, then average RC fill, then average NG fill, then sum and divide by demand for Score (note that fill is fill from demandtrends and NOT just deployed like the field was renamed in 2327)
# Excess is sum of available for each component divided by demand

# In[ ]:


import copy
dmet='demand_met'
emet='excess_met'
#returns % excess demand met for every SRC, AC, phase combination
#used once to find the max and once for actually dataframe computation
def compute_excess(in_df):
    in_df[emet]=(in_df['NG-deployable'] + in_df['AC-deployable'] + in_df['RC-deployable']) / in_df['total-quantity']
    
import numpy as np
#compute % demand met (dmet) and % excess over the demand (emet) 
#first by phase (use average group by with src, ac, phase)
def by_phase_percentages(results_df):
    group_df = results_df.groupby(by=['SRC', 'AC', 'phase']).mean().reset_index()
    #when there is no demand in a phase, dmet is 100%
    group_df[dmet] = np.where((group_df['total-quantity']==0), 1, 
                                                                (group_df['NG-fill'] + 
                                                                group_df['AC-fill'] + 
                                                                group_df['RC-fill']) / group_df['total-quantity'])
    #When there is no demand in a phase, emet is the max emet across all SRCs and phases.
    excess_df = copy.deepcopy(group_df[(group_df['total-quantity'] != 0)])
    compute_excess(excess_df)
    max_excess=excess_df[emet].max()+1
    
    group_df[emet] = np.where((group_df['total-quantity']==0), max_excess, 
                                                        (group_df['NG-deployable'] + 
                                                        group_df['AC-deployable'] + 
                                                        group_df['RC-deployable']) / group_df['total-quantity'])
    #this will be 0 because if there is no demand, we don't have a record.

    return group_df


# Do first: 1 workbook
#     (need to groupby.mean.unstack phase, but what do I expect?)
#     Tab 1: src, ac, results by phase for demand 1, add score, excess
#     Tab 2: src, ac, results by phase in columns for demand 2, add score excess
#     Tab 3: src, ac, score-demand1, excess-demand1, score-dmd2, excess-dmd2, min-demand, min score.

# In[ ]:

def results_by_phase(results_df):
    res=results_df.groupby(by=['SRC', 'AC', 'phase']).mean()
    return res.unstack(level=['phase'])

d_weighted = 'dmet_times_weight'
e_weighted = 'emet_times_weight'
#dmet_sum='weighted_dmet_sum'
#emet_sum='weighted_emet_sum'
dmet_sum=''
emet_sum=''
#given an ordered list of initial columns, put the rest of the columns in the dataframe at the end
def reorder_columns(order, df):
    cols=[c for c in order if c in df] + [c for c in df if c not in order]
    return df[cols]

def ac_not_sorted(group):
    return not all(x>=y for x, y in zip(group[('AC', '')], group[('AC', '')].iloc[1:]))

def add_smoothed(group, col_name, new_name):
    col=[i for i in range(min(group[col_name]), max(group[col_name])+1)]
    col.reverse()
    group[new_name]=col
    return group

def check_order(order_writer, demand_name, left, smooth):
    #Make sure that scores are monotonically decreasing as inventory decreases
    res=left.groupby('SRC', sort=False).filter(ac_not_sorted)
    res[('incorrect_RA', '')]=res[('AC', '')]!=res[('AC_smoothed', '')]
    if res[('incorrect_RA', '')].sum()==0:
        num_flops=0
    else: 
        num_flops=res[('incorrect_RA', '')].sum()-1
    print("There are ", num_flops, " flip flops for ", demand_name)
    res.to_excel(order_writer, sheet_name=demand_name)
    if smooth: 
        left[('AC', '')]=left[('AC_smoothed','')]
    left.drop([('AC_smoothed', '')], axis=1, inplace=True)
    return left

def load_results(results):
    if isinstance(results, pd.DataFrame):        
        #already a dataframe
        df=results
    else:
        df=pd.read_csv(results, sep='\t')
    return df

#compute score and excess from a path to results.txt
def compute_scores(results_path, phase_weights, title_strength, smooth: bool, demand_name, order_writer):
    df=load_results(results_path)
    #sometimes all inventory was equal to 0, but we shouldn't have that. 
    #We should have all phases if all inventory ==0
    df= df[(df[['AC', 'NG', 'RC']] == 0).all(axis=1)==False]
    scores = by_phase_percentages(df)
    scores['weight']=scores['phase'].map(phase_weights)
    scores[d_weighted]=scores[dmet]*scores['weight']
    scores[e_weighted]=scores[emet]*scores['weight']
    res = results_by_phase(scores[['SRC', 'AC', 'NG', 'RC', 
                                   'phase', 'total-quantity', dmet, emet, 
                                   'weight', d_weighted,
                                  e_weighted]])
    res[('Score', dmet_sum)]=res.iloc[:, res.columns.get_level_values(0)==d_weighted].sum(axis=1)
    res[('Excess', emet_sum)]=res.iloc[:, res.columns.get_level_values(0)==e_weighted].sum(axis=1)
    res[('Demand_Total', '')]=res.iloc[:, res.columns.get_level_values(0)=='total-quantity'].sum(axis=1)
    res[('NG_inv', '')]=res.iloc[:, res.columns.get_level_values(0)=='NG'].max(axis=1)
    res[('RC_inv', '')]=res.iloc[:, res.columns.get_level_values(0)=='RC'].max(axis=1)
    res.sort_values(by=[('Score', ''), ('Excess', '')], ascending=False, inplace=True)
    #need to join multindex columns to single index columns in title_strength, so this the merge process
    tuples = [('SRC', ''), ('TITLE', ''), ('STR', '')]
    titles=copy.deepcopy(title_strength)
    titles.columns=pd.MultiIndex.from_tuples(tuples, names=(None, 'phase'))
    res=res.reset_index()
    #Make sure that scores are monotonically decreasing as inventory decreases
    res=res.groupby(('SRC', ''), sort=False)
    #add a smoothed AC column
    res=res.apply(add_smoothed, ('AC', ''), ('AC_smoothed', ''))
    res=check_order(order_writer, demand_name, res, smooth)
    res=res.reset_index(drop=True)
    res = pd.merge(res,
          titles,
          on=[('SRC', '')],
          how='inner'
         )
    res=res.set_index(['SRC', 'AC'])
    res.drop(['NG', 'RC'], axis=1, level=0, inplace=True)
    return res

#The name of the score column used in the combined worksheet before renaming for output
combined_score_out='min_score_peak'
def cols_to_round(df):
    floats = df.select_dtypes('float').columns
    d = dict(zip(floats, [4 for x in floats]))
    d.pop(combined_score_out, None)
    d.pop(('Score', dmet_sum), None)
    return d

#take any multiindex column tuple and if the second level has no value (thus a single entry for both levels),
#then swap the second level with the first level.
def move_col_names_down(df):
    #if the last columns have an empty first level, they will get merge_celled with the previously-titled
    #column, so we use ' ' instead of '' to avoid this.
    new_cols = [(' ', x) if y=='' else (x, y) for (x, y) in df.columns]
    #phase is actually named after the column here
    df.columns=pd.MultiIndex.from_tuples(new_cols, names=(None, 'OML'))

#Find the the most stressful demand by first choosing the lowest total demand, then choosing
#lowest score and then choosing lowest excess
def min_score_demand_total(results_map, row):
    score_excesses = sorted([(-1*row['Demand_Total_'+demand_name], row['Score_'+demand_name], row['Excess_'+demand_name], demand_name) 
                             for demand_name in results_map])
    min_total, min_score, min_excess, min_demand = score_excesses[0]
    return min_demand

def min_score_demand(results_map, row):
    score_excesses = sorted([(row['Score_'+demand_name], row['Excess_'+demand_name], demand_name) 
                             for demand_name in results_map])
    min_score, min_excess, min_demand = score_excesses[0]
    return min_demand

def min_score_demand_peak(peak_map, default_max, row):
    return peak_map.get(row['SRC'], default_max)

#returns the actual minimum score
#pull the score by the demand in the pull column and put it in the out column
def pull_score(df, pull, out):
    df[out]=df.apply(lambda row: row['Score_'+row[pull]], axis=1)

#need to add the excess of the min score case as well for sorting.
def add_excess(row, pull, left):
    score_index = left.columns.get_loc('Score_'+row[pull])
    return row.iloc[score_index+1]

#Given a cell in a sheet starting at row row_start and in column,clear all cell contents
def clear_column(row_start, column, sh):
    for row in range(row_start,sh.max_row):
        if(sh.cell(row,column).value is  None):
            break
        sh.cell(row,column).value= None

# When writing the multi-index dataframes to Excel, pandas put an extra blank row below the column names, which messes up the filter in LibreOffice, but not Excel.  In Excel, you could turn the filter on the blank row.  In LibreOffice, that didn't work.  Although, in LibreOffice, you can turn it on the first row and it captures the first value.  Excel does not.  So those are the filter workarounds, but it looks cleaner to just remove that blank row.
def remove_blank_row(excel_path, out_path):
    if isinstance(excel_path, str):
        wb=openpyxl.reader.excel.load_workbook(excel_path)
    else:
        wb=excel_path
    for sheet in wb.worksheets:
        if not sheet['A3'].value:
            sheet.delete_rows(3, 1)
    wb.save(out_path)

#For the negative scores for last cuts, set the Score to 0 for display and
#do the same with the Excess.
def zero_negative_scores(group_df, excess_col, score_col):
    group_df[excess_col] = np.where((group_df[score_col]<0), 0,                 
                                       group_df[excess_col])
    group_df[score_col] = np.where((group_df[score_col]<0), 0,                 
                                       group_df[score_col])    
    return group_df
    
#add the last ra cuts if there is no rc
def add_last_ra_cuts(scored_results):
    #Since we are using the score from the next lowest inventory, we 
        #need to add additional records to cover the case where we we have no
        #runs from 0 RA 0 NG and 0 RC so that we could cut the entire RA if we
        #wanted to.
        test_frame=scored_results[(scored_results['AC']==1) 
                              & (scored_results['NG_inv']==0) 
                              & (scored_results['RC_inv']==0)].copy()
        test_frame[('Score', '')]=test_frame[('Score', '')]-1.1
        test_frame[('AC', '')]=0
        zero_cols=['demand_met', 'dmet_times_weight', 'emet_times_weight',
                   'excess_met']
        test_frame[zero_cols]=0
        return test_frame
    
#When deciding whether to cut a unit, it makes sense to cut based on what
#the score would be after the cut.  This shifts all scores up one inventory
#level to accomplish that.
def drop_scores(df):
            test_frame=add_last_ra_cuts(df)           
            #add on records to cut the last unit in the inventory.
            scored_results=pd.concat([df, test_frame], ignore_index=True)
            #filter out the base inventories
            scored_results=scored_results[scored_results['AC']!=scored_results['max_AC_inv']]
            #add one to the remaining inventory records
            scored_results['AC']=scored_results['AC']+1
            return scored_results
        
def make_one_n(results_map, peak_max_workbook, out_root, phase_weights, 
               one_n_name, baseline_path, smooth: bool, drop_down: bool=True):
    print("Building ", one_n_name)
    # Read in the SRC baseline for strength and OI title.
    baseline = pd.read_excel(baseline_path)
    title_strength=baseline[['SRC', 'TITLE', 'STR']]
    maxes=pd.read_excel(peak_max_workbook, "by_src")
    maxes['demand_name'] = maxes['demand_name'].astype(str)
    maxes=maxes.set_index('SRC')
    peak_map=maxes['demand_name'].to_dict()
    
    wb = openpyxl.reader.excel.load_workbook(peak_max_workbook)
    ws=wb['default']
    default_max=str(ws['A1'].value)
    
    writer = pd.ExcelWriter(out_root+one_n_name, engine='xlsxwriter')
    left=pd.DataFrame()
    order_path = out_root+"out_of_order.xlsx"
    orders= Path(order_path)
    if orders.is_file():
        order_writer = pd.ExcelWriter(order_path, engine='openpyxl', mode='a', if_sheet_exists='replace')
    else:
        order_writer = pd.ExcelWriter(out_root+"out_of_order.xlsx", engine='xlsxwriter')
            
    for demand_name in results_map:
        #START OF SINGLE DEMAND OUTPUT WORKSHEET
        scored_results = compute_scores(results_map[demand_name], phase_weights, title_strength, smooth, demand_name, order_writer)
        if left.empty:
            max_df=scored_results.reset_index().groupby('SRC')['AC'].apply(max)
            maxes=max_df.to_dict()
        #just to repeat the SRC in the output. Also will add an index on the left.
        scored_results.reset_index(inplace=True)
        #add max ac inventory
        scored_results['max_AC_inv']=scored_results['SRC'].map(maxes)
        if drop_down:
            scored_results=drop_scores(scored_results)    
        #else: 
              #We also get more records now where 0 is the max RA inventory
              #and we had some RC to run.
        #indicate those records that are the base supply
        scored_results['base_supply']=np.where((scored_results['AC']==scored_results['max_AC_inv']), 'X', 'Down')
        #remove maxes
        scored_results.drop(columns=['max_AC_inv'], level=0, inplace=True)
        #TEMP END OF SINGLE DEMAND OUTPUT WORKSHEET
            
        #add scores to all_scores
        #join tables so that you have two score columns for two demands
        #add column called min_score
        #add another column called min_score_demand that indicates which demand this came from
        scored_results=scored_results.set_index(['SRC', 'AC'])
        score_columns=[('Score', dmet_sum), ('Excess', emet_sum), ('Demand_Total', '')]
        score_col_names=['Score_'+demand_name, 'Excess_'+demand_name, 'Demand_Total_'+demand_name]
        scores = scored_results[score_columns]
        scores.columns=score_col_names
        if left.empty:
            left=scores
        else:
            right=scores
            left = pd.merge(left,right,on=['SRC', 'AC'], how='outer')
            
        #RESTART OF SINGLE DEMAND OUTPUT WORKSHEET
        #write to excel file here
        scored_results.reset_index(inplace=True)
        #need to sort again after changing the index..
        scored_results.sort_values(by=[('Score', ''), ('Excess', '')], ascending=False, inplace=True)
        scored_results=zero_negative_scores(scored_results, ('Excess', ''), 
                                            ('Score', ''))
        scored_results.rename(columns={'NG_inv':'NG', 'RC_inv':'AR', 'AC':'RA'}, inplace=True, level=0)
        initial_cols = [('SRC', ''), ('TITLE', ''), ('RA', ''), ('NG', ''), 
                        ('AR', ''),
                       ]
        reordered=reorder_columns(initial_cols, scored_results)
        reordered.reset_index(inplace=True)
        reordered.drop(['index'], axis=1, level=0, inplace=True)
        #avoid rounding when we choose the min score later by doing a deep copy.
        reordered=copy.deepcopy(reordered)
        reordered = reordered.round(cols_to_round(reordered))
        reordered.index=[i for i in range(1, len(reordered.index)+1)]
        move_col_names_down(reordered)
        reordered.rename(columns={"total-quantity": "Demand Days"}, level=0, inplace=True)
        reordered.to_excel(writer, sheet_name=demand_name) 
        #PERMANENT END OF SINGLE DEMAND OUTPUT WORKSHEET
    
    #NOW DOING STUFF TO OUTPUT THE COMBINED SCORE AND EXCESSES THAT WERE COLLECTED
    left.reset_index(inplace=True)
    #deepcopy prevents working on a slice
    left=copy.deepcopy(left)
    left['min_score_demand']=left.apply(partial(min_score_demand, results_map), axis=1)
    left['min_score_demand_total']=left.apply(partial(min_score_demand_total, results_map), axis=1)
    left['min_score_demand_peak']=left.apply(partial(min_score_demand_peak, peak_map, default_max), axis=1)

    #left['min_score']=left.apply(lambda row: row['Score_'+row['min_score_demand']], axis=1)
    pull_score(left, 'min_score_demand', 'min_score')
    pull_score(left, 'min_score_demand_total', 'min_score_total')
    pull_score(left, 'min_score_demand_peak', 'min_score_peak')

    left['min_demand_excess']=left.apply(add_excess, axis=1, pull='min_score_demand', left=left)
    left['min_demand_total_excess']=left.apply(add_excess, axis=1, pull='min_score_demand_total', left=left)
    left['min_demand_peak_excess']=left.apply(add_excess, axis=1, pull='min_score_demand_peak', left=left)
    #write third worksheet with all scores here
    left = pd.merge(left, title_strength, on='SRC', validate='many_to_one')
    left['SRC2']=left.SRC.str[:2]
    initial_cols=['SRC2','SRC', 'TITLE', 'AC']
    left = reorder_columns(initial_cols, left)
    #resort after choosing the max demand records from each demand case.
    left.sort_values(by=['min_score_peak', 'min_demand_peak_excess'], ascending=False, inplace=True)
    left.reset_index(inplace=True)
    left.drop(['index'], axis=1, inplace=True)
    left = left.round(cols_to_round(left))
    
    #only select final output here
    left=left[initial_cols + ['min_score_demand_peak', 'min_score_peak', 'min_demand_peak_excess', 'STR']]
    left.columns=['SRC2', 'SRC', 'TITLE', 'RA Qty', 'Most Stressful', 'Score', 'Excess', 'STR']
    left.index=[i for i in range(1, len(left.index)+1)]  
    left=zero_negative_scores(left, 'Excess', 'Score')
     #output     
    left.to_excel(writer, sheet_name='combined')
    writer.close()
    order_writer.close()
    wb = openpyxl.reader.excel.load_workbook(out_root+one_n_name)
    ws=wb['combined']
    ws.cell(1, 1).value = "OML"
    remove_blank_row(wb, out_root+one_n_name)
    
def concat_run_to_phase(run_key, results_path):
    df=pd.read_csv(results_path, sep='\t')
    df['phase']=df['phase']+'-'+run_key
    return df

#Before we call make_one_n, concatenate all of the results files into one
#DataFrame, concatenating phase labels with the run key in the results_map.
def one_n_across_runs(results_map, peak_max_workbook, out_root, phase_weights, one_n_name, baseline_path, smooth: bool):
    concatenated_dfs=[]
    for demand_name in results_map:
        concatenated_df=concat_run_to_phase(demand_name, results_map[demand_name])
        concatenated_dfs.append(concatenated_df)
    all_runs_df=pd.concat(concatenated_dfs)
    dummy_results_map={'all_runs' : all_runs_df}
    make_one_n(dummy_results_map, peak_max_workbook, out_root, phase_weights, one_n_name, baseline_path, smooth)
    
#So far, we've had an agreed upon weighting for each phase of the demand, and
#now that we're weighting results across multiple Marathon runs, we want to say
#demand c has 75% weight and demand A has 25% weight and then multiple those
#weights by the weight for each phase.
#results_weight looks like {demand_A : .25, demand_C : .75}
#phase_breakout looks like {comp: 75, phase_1 : .06, phase_2: .19}
#there might be cases where we don't want to concat the run name to the phase
#like if we only have one entry in the results_map.
def split_run_weights(results_weights, phase_breakout, 
                      concat_run: bool=True):
    all_weights={}
    for demand_name in results_weights:
        for phase in phase_breakout:
            if concat_run:
                k=phase + '-' + demand_name
            else:
                k=phase
            all_weights[k]= phase_breakout[phase]*results_weights[demand_name]
    return all_weights
        
        
        

    







